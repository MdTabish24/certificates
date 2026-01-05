import os
import qrcode
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from PyPDF2 import PdfReader, PdfWriter
import io
import base64
import httpx
import json
from dotenv import load_dotenv

# Load .env file
load_dotenv()

# Configuration
EXCEL_FILE = "CERTIFIED STUDENTS (1).xlsx"
TEMPLATE_PDF = "Usdc Certificate.pdf"

# GitHub Pages URL for verification (no sleep!)
GITHUB_USERNAME = "MdTabish24"
REPO_NAME = "certificates"
VERIFY_BASE_URL = f"https://{GITHUB_USERNAME}.github.io/{REPO_NAME}"

# ImageKit Config - Environment variable se
IMAGEKIT_PRIVATE_KEY = os.environ.get('IMAGEKIT_PRIVATE_KEY', '')

DATA_FILE = "certificates_data.json"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_data(cert_id, data):
    all_data = load_data()
    all_data[cert_id] = data
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

def upload_to_imagekit(file_buffer, filename, folder):
    try:
        file_buffer.seek(0)
        file_base64 = base64.b64encode(file_buffer.read()).decode('utf-8')
        url = "https://upload.imagekit.io/api/v1/files/upload"
        data = {'file': file_base64, 'fileName': filename, 'folder': folder}
        auth = (IMAGEKIT_PRIVATE_KEY, '')
        response = httpx.post(url, data=data, auth=auth, timeout=60)
        if response.status_code == 200:
            print(f"    [UPLOAD] {folder}/{filename}")
            return response.json()
        else:
            print(f"    [ERROR] {response.status_code}")
            return None
    except Exception as e:
        print(f"    [ERROR] {e}")
        return None

def is_green(row_cells):
    greens = ['00FF00','92D050','00B050','C6EFCE','98FB98','90EE90']
    for cell in row_cells:
        if cell.fill and cell.fill.fgColor:
            c = cell.fill.fgColor.rgb
            if c and isinstance(c, str):
                if len(c) == 8: c = c[2:]
                if c.upper() in [g.upper() for g in greens]:
                    return True
    return False

def main():
    if not IMAGEKIT_PRIVATE_KEY:
        print("[ERROR] IMAGEKIT_PRIVATE_KEY environment variable set karo!")
        print("\nWindows CMD:")
        print('  set IMAGEKIT_PRIVATE_KEY=private_xxxxx')
        print("\nPowerShell:")
        print('  $env:IMAGEKIT_PRIVATE_KEY="private_xxxxx"')
        return
    
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    print(f"Excel: {EXCEL_FILE}, Rows: {sheet.max_row}")
    
    count = 0
    for row_idx in range(2, sheet.max_row + 1):
        row_cells = [sheet.cell(row=row_idx, column=col) for col in range(1, sheet.max_column + 1)]
        if is_green(row_cells):
            continue
        
        sr = sheet.cell(row=row_idx, column=1).value
        name = sheet.cell(row=row_idx, column=2).value
        grade = sheet.cell(row=row_idx, column=3).value
        course = sheet.cell(row=row_idx, column=4).value
        aadhaar = sheet.cell(row=row_idx, column=5).value
        doi = sheet.cell(row=row_idx, column=7).value
        duration = sheet.cell(row=row_idx, column=9).value
        
        if not name: continue
        
        doi_str = doi.strftime('%d-%m-%Y') if doi else "N/A"
        cert_id = f"CERT{sr:04d}" if isinstance(sr, int) else f"CERT{row_idx:04d}"
        
        print(f"\n[{count+1}] {cert_id}: {name}")
        
        # Create HTML verification page for GitHub Pages
        os.makedirs("docs", exist_ok=True)
        html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Verify {cert_id}</title>
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:Segoe UI,sans-serif;background:linear-gradient(135deg,#667eea,#764ba2);min-height:100vh;display:flex;justify-content:center;align-items:center;padding:20px}}.card{{background:#fff;border-radius:20px;box-shadow:0 20px 60px rgba(0,0,0,.3);max-width:500px;width:100%;padding:40px}}.header{{text-align:center;margin-bottom:30px}}.header h1{{color:#667eea;font-size:28px}}.header p{{color:#666}}.info{{margin:15px 0;padding:15px;background:#f8f9fa;border-radius:10px;border-left:4px solid #667eea}}.label{{color:#666;font-size:12px;text-transform:uppercase}}.value{{color:#333;font-size:18px;font-weight:600}}.badge{{background:#10b981;color:#fff;padding:15px;border-radius:50px;text-align:center;margin-top:25px;font-weight:600}}</style>
</head><body><div class="card"><div class="header"><h1>✅ Verified</h1><p>{cert_id}</p></div>
<div class="info"><div class="label">Name</div><div class="value">{name}</div></div>
<div class="info"><div class="label">Course</div><div class="value">{course or 'N/A'}</div></div>
<div class="info"><div class="label">Aadhaar</div><div class="value">{aadhaar or 'N/A'}</div></div>
<div class="badge">✓ Verified & Authentic</div></div></body></html>'''
        
        with open(f"docs/{cert_id}.html", "w", encoding="utf-8") as f:
            f.write(html)
        
        # QR Code - GitHub Pages URL (no sleep!)
        verify_url = f"{VERIFY_BASE_URL}/{cert_id}.html"
        qr = qrcode.make(verify_url)
        qr_buf = io.BytesIO()
        qr.save(qr_buf, format='PNG')
        qr_buf.seek(0)
        
        # PDF
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=landscape(A4))
        can.setFont("Times-BoldItalic", 16)
        
        pw = landscape(A4)[0]
        nw = can.stringWidth(str(name), "Times-BoldItalic", 16)
        can.drawString((pw-nw)/2, 330, str(name))
        can.drawString(375, 263, str(course) if course else "")
        can.drawString(230, 210, str(duration) if duration else "")
        can.drawString(187, 170, str(aadhaar) if aadhaar else "")
        can.drawString(680, 167, str(grade) if grade else "")
        can.drawString(470, 167, doi_str)
        
        # QR on PDF
        tmp = f"tmp_{cert_id}.png"
        with open(tmp, 'wb') as f: f.write(qr_buf.getvalue())
        can.drawImage(tmp, 87, 87, width=60, height=60)
        os.remove(tmp)
        
        can.save()
        packet.seek(0)
        
        template = PdfReader(TEMPLATE_PDF)
        overlay = PdfReader(packet)
        output = PdfWriter()
        page = template.pages[0]
        page.merge_page(overlay.pages[0])
        output.add_page(page)
        
        pdf_buf = io.BytesIO()
        output.write(pdf_buf)
        pdf_buf.seek(0)
        
        # Upload
        safe = "".join(c for c in str(name) if c.isalnum() or c in ' -_').strip()
        pdf_res = upload_to_imagekit(pdf_buf, f"{cert_id}_{safe}.pdf", "/Generated_Certificates")
        qr_buf.seek(0)
        qr_res = upload_to_imagekit(qr_buf, f"{cert_id}.png", "/QR_Codes")
        
        save_data(cert_id, {
            'name': str(name), 'course': str(course or ''),
            'aadhaar': str(aadhaar or ''), 'doi': doi_str,
            'duration': str(duration or ''), 'grade': str(grade or ''),
            'pdf_url': pdf_res.get('url','') if pdf_res else '',
            'qr_url': qr_res.get('url','') if qr_res else ''
        })
        count += 1
    
    print(f"\n[DONE] {count} certificates generated!")
    print(f"\n[NEXT] GitHub pe HTML push kar:")
    print("  git add docs/*.html")
    print("  git commit -m 'Add verification pages'")
    print("  git push origin main")
    print(f"\n[URL] QR scan karega toh: {VERIFY_BASE_URL}/CERT0001.html")

if __name__ == "__main__":
    main()
